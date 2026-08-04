"""
ΣΑΠ-ΦΘ
Unified Final Report
"""

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from import_data import load_data
from metadata_engine import build_metadata
from overall_iqi import calculate_iqi
from scoring_engine import build_scored_dataset
from section_scores import section_means

from question_analytics import (
    calculate_question_statistics,
    create_priority_tables,
    create_section_summary,
)

from text_analytics import (
    build_all_keywords_table,
    build_comments_table,
    build_keyword_table,
    build_question_summary,
)


PROJECT_ROOT = Path(__file__).resolve().parent.parent

OUTPUT_FILE = (
    PROJECT_ROOT
    / "excel"
    / "ΣΑΠ-ΦΘ_FINAL_REPORT.xlsx"
)


SECTION_ORDER = [
    "Α",
    "Β",
    "Γ",
    "Δ",
    "Ε",
    "ΣΤ",
    "Ζ",
    "Η",
    "Θ",
]


SECTION_TITLES = {
    "Α": "Υποδοχή και ενημέρωση",
    "Β": "Οργάνωση πρακτικής άσκησης",
    "Γ": "Κλινική εκπαίδευση και εμπειρία",
    "Δ": "Υποδομές και εξοπλισμός",
    "Ε": "Συνεργασία με επαγγελματίες υγείας",
    "ΣΤ": "Καθοδήγηση και υποστήριξη",
    "Ζ": "Ηγεσία και επαγγελματισμός",
    "Η": "Κλίμα συνεργασίας και ασφάλειας",
    "Θ": "Χώροι και συνθήκες πρακτικής",
}


TITLE_FILL = PatternFill(
    fill_type="solid",
    fgColor="1F4E78",
)

SUBTITLE_FILL = PatternFill(
    fill_type="solid",
    fgColor="D9EAF7",
)

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="5B9BD5",
)

SUCCESS_FILL = PatternFill(
    fill_type="solid",
    fgColor="E2F0D9",
)

WARNING_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFF2CC",
)

WHITE_FONT = Font(
    color="FFFFFF",
    bold=True,
)

THIN_BORDER = Border(
    left=Side(
        style="thin",
        color="B7B7B7",
    ),
    right=Side(
        style="thin",
        color="B7B7B7",
    ),
    top=Side(
        style="thin",
        color="B7B7B7",
    ),
    bottom=Side(
        style="thin",
        color="B7B7B7",
    ),
)


def excel_safe_value(value):
    """
    Μετατρέπει τιμές pandas σε τιμές συμβατές με Excel.
    """

    if value is None:
        return None

    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass

    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()

    return value


def section_sort_value(section):
    """
    Επιστρέφει τη σειρά εμφάνισης μιας ενότητας.
    """

    if section in SECTION_ORDER:
        return SECTION_ORDER.index(section)

    return len(SECTION_ORDER)


def prepare_section_scores(scores):
    """
    Προσθέτει περιγραφή και ταξινομεί τις ενότητες.
    """

    if scores.empty:
        return pd.DataFrame(
            columns=[
                "Ενότητα",
                "Περιγραφή ενότητας",
                "Μέσος όρος",
            ]
        )

    prepared = scores.copy()

    prepared["Section"] = (
        prepared["Section"]
        .astype(str)
    )

    prepared["_order"] = (
        prepared["Section"]
        .apply(section_sort_value)
    )

    prepared["Περιγραφή ενότητας"] = (
        prepared["Section"]
        .map(SECTION_TITLES)
        .fillna("")
    )

    prepared = prepared.sort_values(
        by=[
            "_order",
            "Section",
        ]
    )

    prepared = prepared.rename(
        columns={
            "Section": "Ενότητα",
            "Mean": "Μέσος όρος",
        }
    )

    return prepared[
        [
            "Ενότητα",
            "Περιγραφή ενότητας",
            "Μέσος όρος",
        ]
    ].reset_index(drop=True)


def write_dataframe(
    worksheet,
    dataframe,
    start_row=1,
    start_column=1,
    include_header=True,
):
    """
    Γράφει DataFrame σε φύλλο Excel.
    """

    rows = dataframe_to_rows(
        dataframe,
        index=False,
        header=include_header,
    )

    for row_offset, row_values in enumerate(rows):
        current_row = start_row + row_offset

        for column_offset, value in enumerate(row_values):
            current_column = start_column + column_offset

            worksheet.cell(
                row=current_row,
                column=current_column,
                value=excel_safe_value(value),
            )


def style_header(
    worksheet,
    row_number,
    first_column,
    last_column,
):
    """
    Μορφοποιεί γραμμή επικεφαλίδων.
    """

    for column_number in range(
        first_column,
        last_column + 1,
    ):
        cell = worksheet.cell(
            row=row_number,
            column=column_number,
        )

        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = THIN_BORDER

    worksheet.row_dimensions[
        row_number
    ].height = 35


def apply_borders(
    worksheet,
    first_row,
    last_row,
    first_column,
    last_column,
):
    """
    Προσθέτει περιγράμματα σε περιοχή κελιών.
    """

    for row_number in range(
        first_row,
        last_row + 1,
    ):
        for column_number in range(
            first_column,
            last_column + 1,
        ):
            worksheet.cell(
                row=row_number,
                column=column_number,
            ).border = THIN_BORDER


def auto_fit_columns(
    worksheet,
    minimum_width=12,
    maximum_width=65,
):
    """
    Προσαρμόζει τα πλάτη των στηλών.
    """

    for column_number in range(
        1,
        worksheet.max_column + 1,
    ):
        column_letter = get_column_letter(
            column_number
        )

        maximum_length = 0

        for row_number in range(
            1,
            worksheet.max_row + 1,
        ):
            value = worksheet.cell(
                row=row_number,
                column=column_number,
            ).value

            if value is None:
                continue

            maximum_length = max(
                maximum_length,
                len(str(value)),
            )

        worksheet.column_dimensions[
            column_letter
        ].width = min(
            max(
                maximum_length + 3,
                minimum_width,
            ),
            maximum_width,
        )


def style_data_sheet(
    worksheet,
    wrap_columns=None,
    percentage_columns=None,
    decimal_columns=None,
):
    """
    Εφαρμόζει κοινή μορφοποίηση σε φύλλο δεδομένων.
    """

    wrap_columns = wrap_columns or []
    percentage_columns = percentage_columns or []
    decimal_columns = decimal_columns or []

    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_view.zoomScale = 85

    if worksheet.max_column:
        style_header(
            worksheet=worksheet,
            row_number=1,
            first_column=1,
            last_column=worksheet.max_column,
        )

        worksheet.auto_filter.ref = (
            f"A1:"
            f"{get_column_letter(worksheet.max_column)}"
            f"{worksheet.max_row}"
        )

    apply_borders(
        worksheet=worksheet,
        first_row=1,
        last_row=worksheet.max_row,
        first_column=1,
        last_column=worksheet.max_column,
    )

    for row_number in range(
        2,
        worksheet.max_row + 1,
    ):
        worksheet.row_dimensions[
            row_number
        ].height = 38

        for column_number in range(
            1,
            worksheet.max_column + 1,
        ):
            cell = worksheet.cell(
                row=row_number,
                column=column_number,
            )

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=(
                    column_number
                    in wrap_columns
                ),
            )

            if column_number in percentage_columns:
                cell.number_format = "0.00"

            if column_number in decimal_columns:
                cell.number_format = "0.00"

    auto_fit_columns(worksheet)


def create_executive_summary(
    workbook,
    raw_data,
    section_scores,
    question_statistics,
    comments,
    iqi_value,
):
    """
    Δημιουργεί την κεντρική διοικητική σύνοψη.
    """

    worksheet = workbook.active
    worksheet.title = "Executive_Summary"

    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_view.zoomScale = 85

    worksheet.merge_cells("A1:J1")
    worksheet["A1"] = "ΣΑΠ-ΦΘ"

    worksheet["A1"].font = Font(
        size=22,
        bold=True,
        color="FFFFFF",
    )

    worksheet["A1"].fill = TITLE_FILL

    worksheet["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    worksheet.row_dimensions[1].height = 38

    worksheet.merge_cells("A2:J2")

    worksheet["A2"] = (
        "Ενοποιημένη Αναφορά Αξιολόγησης "
        "Πρακτικής Άσκησης Φυσικοθεραπείας"
    )

    worksheet["A2"].font = Font(
        size=13,
        bold=True,
    )

    worksheet["A2"].fill = SUBTITLE_FILL

    worksheet["A2"].alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )

    worksheet.row_dimensions[2].height = 32

    worksheet["A4"] = "Συνολικός δείκτης IQI"
    worksheet["A4"].font = Font(
        bold=True,
        size=12,
    )
    worksheet["A4"].border = THIN_BORDER
    worksheet["A4"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    worksheet["B4"] = excel_safe_value(iqi_value)
    worksheet["B4"].font = Font(
        bold=True,
        size=20,
        color="1F4E78",
    )
    worksheet["B4"].fill = SUCCESS_FILL
    worksheet["B4"].border = THIN_BORDER
    worksheet["B4"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )
    worksheet["B4"].number_format = "0.00"

    worksheet["D4"] = "Αριθμός φοιτητών"
    worksheet["D4"].font = Font(
        bold=True,
        size=12,
    )
    worksheet["D4"].border = THIN_BORDER
    worksheet["D4"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    worksheet["E4"] = len(raw_data)
    worksheet["E4"].font = Font(
        bold=True,
        size=18,
        color="1F4E78",
    )
    worksheet["E4"].fill = SUCCESS_FILL
    worksheet["E4"].border = THIN_BORDER
    worksheet["E4"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    worksheet["G4"] = "Ερωτήσεις Likert"
    worksheet["G4"].font = Font(
        bold=True,
        size=12,
    )
    worksheet["G4"].border = THIN_BORDER
    worksheet["G4"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    worksheet["H4"] = len(question_statistics)
    worksheet["H4"].font = Font(
        bold=True,
        size=18,
        color="1F4E78",
    )
    worksheet["H4"].fill = SUCCESS_FILL
    worksheet["H4"].border = THIN_BORDER
    worksheet["H4"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    worksheet["A6"] = "Συνολικά σχόλια"
    worksheet["A6"].font = Font(
        bold=True,
    )
    worksheet["A6"].border = THIN_BORDER
    worksheet["A6"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    worksheet["B6"] = len(comments)
    worksheet["B6"].font = Font(
        bold=True,
        size=16,
        color="1F4E78",
    )
    worksheet["B6"].fill = SUCCESS_FILL
    worksheet["B6"].border = THIN_BORDER
    worksheet["B6"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    worksheet["D6"] = "Ενότητες"
    worksheet["D6"].font = Font(
        bold=True,
    )
    worksheet["D6"].border = THIN_BORDER
    worksheet["D6"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    worksheet["E6"] = len(section_scores)
    worksheet["E6"].font = Font(
        bold=True,
        size=16,
        color="1F4E78",
    )
    worksheet["E6"].fill = SUCCESS_FILL
    worksheet["E6"].border = THIN_BORDER
    worksheet["E6"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    worksheet["G6"] = "Ημερομηνία αναφοράς"
    worksheet["G6"].font = Font(
        bold=True,
    )
    worksheet["G6"].border = THIN_BORDER
    worksheet["G6"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    worksheet["H6"] = datetime.now()
    worksheet["H6"].number_format = "dd/mm/yyyy hh:mm"
    worksheet["H6"].fill = SUCCESS_FILL
    worksheet["H6"].border = THIN_BORDER
    worksheet["H6"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    worksheet["A8"] = "Ενότητα"
    worksheet["B8"] = "Περιγραφή ενότητας"
    worksheet["C8"] = "Μέσος όρος"

    style_header(
        worksheet=worksheet,
        row_number=8,
        first_column=1,
        last_column=3,
    )

    start_row = 9

    for row_offset, record in section_scores.iterrows():
        current_row = start_row + row_offset

        worksheet.cell(
            row=current_row,
            column=1,
            value=record["Ενότητα"],
        )

        worksheet.cell(
            row=current_row,
            column=2,
            value=record["Περιγραφή ενότητας"],
        )

        worksheet.cell(
            row=current_row,
            column=3,
            value=excel_safe_value(
                record["Μέσος όρος"]
            ),
        )

        worksheet.cell(
            row=current_row,
            column=1,
        ).alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        worksheet.cell(
            row=current_row,
            column=2,
        ).alignment = Alignment(
            vertical="center",
            wrap_text=True,
        )

        worksheet.cell(
            row=current_row,
            column=3,
        ).alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        worksheet.cell(
            row=current_row,
            column=3,
        ).number_format = "0.00"

        worksheet.row_dimensions[
            current_row
        ].height = 34

    end_row = (
        start_row + len(section_scores) - 1
        if len(section_scores)
        else start_row
    )

    apply_borders(
        worksheet=worksheet,
        first_row=8,
        last_row=end_row,
        first_column=1,
        last_column=3,
    )

    if len(section_scores):
        worksheet.conditional_formatting.add(
            f"C{start_row}:C{end_row}",
            ColorScaleRule(
                start_type="num",
                start_value=1,
                start_color="F8696B",
                mid_type="num",
                mid_value=3,
                mid_color="FFEB84",
                end_type="num",
                end_value=5,
                end_color="63BE7B",
            ),
        )

        chart = BarChart()

        chart.type = "col"
        chart.grouping = "clustered"
        chart.gapWidth = 65

        chart.title = "Αξιολόγηση ανά Ενότητα"
        chart.legend = None

        chart.height = 13
        chart.width = 22

        chart.y_axis.title = None
        chart.x_axis.title = None

        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 5
        chart.y_axis.majorUnit = 1

        chart.x_axis.delete = False
        chart.x_axis.tickLblPos = "nextTo"
        chart.x_axis.tickLblSkip = 1

        data_reference = Reference(
            worksheet,
            min_col=3,
            min_row=start_row,
            max_row=end_row,
        )

        category_reference = Reference(
            worksheet,
            min_col=1,
            min_row=start_row,
            max_row=end_row,
        )

        chart.add_data(
            data_reference,
            titles_from_data=False,
        )

        chart.set_categories(
            category_reference
        )

        chart.dLbls = DataLabelList()
        chart.dLbls.showVal = True
        chart.dLbls.showCatName = False
        chart.dLbls.showSerName = False
        chart.dLbls.showLegendKey = False
        chart.dLbls.position = "outEnd"
        chart.dLbls.numFmt = "0.00"

        worksheet.add_chart(
            chart,
            "E8",
        )

    worksheet.column_dimensions["A"].width = 25
    worksheet.column_dimensions["B"].width = 52
    worksheet.column_dimensions["C"].width = 18
    worksheet.column_dimensions["D"].width = 22

    for column_letter in [
        "E",
        "F",
        "G",
        "H",
        "I",
        "J",
    ]:
        worksheet.column_dimensions[
            column_letter
        ].width = 16


def add_dataframe_sheet(
    workbook,
    title,
    dataframe,
    wrap_columns=None,
    percentage_columns=None,
    decimal_columns=None,
):
    """
    Προσθέτει φύλλο δεδομένων στην αναφορά.
    """

    worksheet = workbook.create_sheet(
        title=title
    )

    write_dataframe(
        worksheet=worksheet,
        dataframe=dataframe,
    )

    style_data_sheet(
        worksheet=worksheet,
        wrap_columns=wrap_columns,
        percentage_columns=percentage_columns,
        decimal_columns=decimal_columns,
    )

    return worksheet


def create_final_report():
    """
    Δημιουργεί το ενοποιημένο αρχείο τελικής αναφοράς.
    """

    raw_data = load_data()

    scored_data = build_scored_dataset()

    metadata = build_metadata()

    iqi_value = calculate_iqi()

    section_scores = prepare_section_scores(
        section_means()
    )

    question_statistics = (
        calculate_question_statistics(
            raw_data=raw_data,
            scored_data=scored_data,
            metadata=metadata,
        )
    )

    section_summary = create_section_summary(
        question_statistics
    )

    strongest_points, improvement_priorities = (
        create_priority_tables(
            question_statistics
        )
    )

    comments = build_comments_table(
        raw_data=raw_data,
        metadata=metadata,
    )

    text_summary = build_question_summary(
        raw_data=raw_data,
        metadata=metadata,
    )

    keywords_by_question = build_keyword_table(
        comments
    )

    overall_keywords = build_all_keywords_table(
        comments
    )

    workbook = Workbook()

    create_executive_summary(
        workbook=workbook,
        raw_data=raw_data,
        section_scores=section_scores,
        question_statistics=question_statistics,
        comments=comments,
        iqi_value=iqi_value,
    )

    add_dataframe_sheet(
        workbook=workbook,
        title="Section_Summary",
        dataframe=section_summary,
        wrap_columns=[2],
        percentage_columns=[
            12,
            13,
            14,
        ],
        decimal_columns=[7],
    )

    question_sheet = add_dataframe_sheet(
        workbook=workbook,
        title="Question_Statistics",
        dataframe=question_statistics,
        wrap_columns=[
            2,
            3,
        ],
        percentage_columns=[
            9,
            18,
            19,
            20,
        ],
        decimal_columns=[
            10,
            11,
            12,
            13,
            14,
        ],
    )

    if (
        not question_statistics.empty
        and "Μέσος όρος"
        in question_statistics.columns
    ):
        mean_column = (
            question_statistics.columns.get_loc(
                "Μέσος όρος"
            )
            + 1
        )

        mean_letter = get_column_letter(
            mean_column
        )

        question_sheet.conditional_formatting.add(
            f"{mean_letter}2:"
            f"{mean_letter}{question_sheet.max_row}",
            ColorScaleRule(
                start_type="num",
                start_value=1,
                start_color="F8696B",
                mid_type="num",
                mid_value=3,
                mid_color="FFEB84",
                end_type="num",
                end_value=5,
                end_color="63BE7B",
            ),
        )

    add_dataframe_sheet(
        workbook=workbook,
        title="Strongest_Points",
        dataframe=strongest_points,
        wrap_columns=[2],
        percentage_columns=[
            5,
            6,
        ],
        decimal_columns=[4],
    )

    improvement_sheet = add_dataframe_sheet(
        workbook=workbook,
        title="Improvement_Priorities",
        dataframe=improvement_priorities,
        wrap_columns=[2],
        percentage_columns=[
            5,
            6,
        ],
        decimal_columns=[4],
    )

    for row_number in range(
        2,
        improvement_sheet.max_row + 1,
    ):
        improvement_sheet.cell(
            row=row_number,
            column=4,
        ).fill = WARNING_FILL

    add_dataframe_sheet(
        workbook=workbook,
        title="Text_Summary",
        dataframe=text_summary,
        wrap_columns=[1],
        percentage_columns=[5],
        decimal_columns=[
            6,
            7,
        ],
    )

    add_dataframe_sheet(
        workbook=workbook,
        title="All_Comments",
        dataframe=comments,
        wrap_columns=[
            2,
            3,
        ],
    )

    add_dataframe_sheet(
        workbook=workbook,
        title="Keywords_By_Question",
        dataframe=keywords_by_question,
        wrap_columns=[1],
    )

    add_dataframe_sheet(
        workbook=workbook,
        title="Overall_Keywords",
        dataframe=overall_keywords,
    )

    OUTPUT_FILE.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook.save(OUTPUT_FILE)

    print(
        f"Τελική αναφορά δημιουργήθηκε: {OUTPUT_FILE}"
    )

    print(
        f"Φοιτητές: {len(raw_data)}"
    )

    print(
        "Ερωτήσεις Likert: "
        f"{len(question_statistics)}"
    )

    print(
        f"Ανοιχτά σχόλια: {len(comments)}"
    )


if __name__ == "__main__":
    create_final_report()