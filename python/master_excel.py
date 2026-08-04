"""
ΣΑΠ-ΦΘ
MASTER Excel Builder
"""

from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from config import MASTER_FILE
from import_data import load_data
from metadata_engine import build_metadata
from overall_iqi import calculate_iqi
from scoring_engine import build_scored_dataset
from section_scores import section_means


SECTION_ORDER = [
    "Α",
    "Β",
    "Γ",
    "Δ",
    "Ε",
    "ΣΤ",
    "Ζ",
    "Η",
    "Θ",
]


SECTION_TITLES = {
    "Α": "Υποδοχή και ενημέρωση",
    "Β": "Οργάνωση πρακτικής άσκησης",
    "Γ": "Κλινική εκπαίδευση και εμπειρία",
    "Δ": "Υποδομές και εξοπλισμός",
    "Ε": "Συνεργασία με επαγγελματίες υγείας",
    "ΣΤ": "Καθοδήγηση και υποστήριξη",
    "Ζ": "Ηγεσία και επαγγελματισμός",
    "Η": "Κλίμα συνεργασίας και ασφάλειας",
    "Θ": "Χώροι και συνθήκες πρακτικής",
}


TITLE_FILL = PatternFill(
    fill_type="solid",
    fgColor="1F4E78",
)

SUBTITLE_FILL = PatternFill(
    fill_type="solid",
    fgColor="D9EAF7",
)

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="5B9BD5",
)

IQI_FILL = PatternFill(
    fill_type="solid",
    fgColor="E2F0D9",
)

WHITE_FONT = Font(
    color="FFFFFF",
    bold=True,
)

THIN_BORDER = Border(
    left=Side(style="thin", color="B7B7B7"),
    right=Side(style="thin", color="B7B7B7"),
    top=Side(style="thin", color="B7B7B7"),
    bottom=Side(style="thin", color="B7B7B7"),
)


def excel_safe_value(value):
    """
    Μετατρέπει τιμές pandas σε μορφή συμβατή με το Excel.
    """

    if value is None:
        return None

    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass

    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()

    return value


def prepare_section_scores(scores):
    """
    Ταξινομεί τις ενότητες και προσθέτει περιγραφή.
    """

    if scores.empty:
        return pd.DataFrame(
            columns=[
                "Section",
                "Description",
                "Mean",
            ]
        )

    prepared = scores.copy()

    prepared["Section"] = prepared["Section"].astype(str)

    prepared["Order"] = prepared["Section"].apply(
        lambda section: (
            SECTION_ORDER.index(section)
            if section in SECTION_ORDER
            else len(SECTION_ORDER)
        )
    )

    prepared["Description"] = prepared["Section"].apply(
        lambda section: SECTION_TITLES.get(
            section,
            "Μη καθορισμένη ενότητα",
        )
    )

    prepared = prepared.sort_values(
        by=[
            "Order",
            "Section",
        ]
    )

    return prepared[
        [
            "Section",
            "Description",
            "Mean",
        ]
    ].reset_index(drop=True)


def write_dataframe(
    worksheet,
    dataframe,
    start_row=1,
    start_column=1,
    include_header=True,
):
    """
    Γράφει ένα pandas DataFrame σε φύλλο Excel.
    """

    rows = dataframe_to_rows(
        dataframe,
        index=False,
        header=include_header,
    )

    for row_offset, row_values in enumerate(rows):
        current_row = start_row + row_offset

        for column_offset, value in enumerate(row_values):
            current_column = start_column + column_offset

            worksheet.cell(
                row=current_row,
                column=current_column,
                value=excel_safe_value(value),
            )


def style_table_header(
    worksheet,
    row_number,
    first_column,
    last_column,
):
    """
    Μορφοποιεί μία γραμμή ως επικεφαλίδα πίνακα.
    """

    for column_number in range(
        first_column,
        last_column + 1,
    ):
        cell = worksheet.cell(
            row=row_number,
            column=column_number,
        )

        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = THIN_BORDER


def apply_table_borders(
    worksheet,
    first_row,
    last_row,
    first_column,
    last_column,
):
    """
    Προσθέτει περιγράμματα σε περιοχή κελιών.
    """

    for row_number in range(
        first_row,
        last_row + 1,
    ):
        for column_number in range(
            first_column,
            last_column + 1,
        ):
            worksheet.cell(
                row=row_number,
                column=column_number,
            ).border = THIN_BORDER


def auto_fit_columns(
    worksheet,
    minimum_width=10,
    maximum_width=70,
):
    """
    Προσαρμόζει αυτόματα το πλάτος των στηλών.
    """

    for column_cells in worksheet.columns:
        column_letter = get_column_letter(
            column_cells[0].column
        )

        maximum_length = 0

        for cell in column_cells:
            if cell.value is None:
                continue

            value_length = len(str(cell.value))

            if value_length > maximum_length:
                maximum_length = value_length

        calculated_width = maximum_length + 3

        worksheet.column_dimensions[
            column_letter
        ].width = min(
            max(calculated_width, minimum_width),
            maximum_width,
        )


def create_dashboard(
    workbook,
    raw_data,
    scores,
    iqi_value,
):
    """
    Δημιουργεί το κεντρικό Dashboard.
    """

    worksheet = workbook.active
    worksheet.title = "Dashboard"

    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_view.zoomScale = 85
    worksheet.freeze_panes = None

    worksheet.merge_cells("A1:M1")
    worksheet["A1"] = "ΣΑΠ-ΦΘ"

    worksheet["A1"].font = Font(
        size=22,
        bold=True,
        color="FFFFFF",
    )

    worksheet["A1"].fill = TITLE_FILL

    worksheet["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    worksheet.row_dimensions[1].height = 36

    worksheet.merge_cells("A2:M2")
    worksheet["A2"] = (
        "Σύστημα Αξιολόγησης και Ποιότητας "
        "Πρακτικής Άσκησης Φυσικοθεραπείας"
    )

    worksheet["A2"].font = Font(
        size=12,
        bold=True,
        color="1F1F1F",
    )

    worksheet["A2"].fill = SUBTITLE_FILL

    worksheet["A2"].alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )

    worksheet.row_dimensions[2].height = 30

    # Κεντρικός δείκτης IQI.
    worksheet["A4"] = "Συνολικός Δείκτης IQI"
    worksheet["A4"].font = Font(
        size=13,
        bold=True,
    )
    worksheet["A4"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )
    worksheet["A4"].border = THIN_BORDER

    worksheet["B4"] = excel_safe_value(iqi_value)
    worksheet["B4"].font = Font(
        size=20,
        bold=True,
        color="1F4E78",
    )
    worksheet["B4"].fill = IQI_FILL
    worksheet["B4"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )
    worksheet["B4"].number_format = "0.00"
    worksheet["B4"].border = THIN_BORDER

    # Αριθμός απαντήσεων.
    worksheet["D4"] = "Αριθμός απαντήσεων"
    worksheet["D4"].font = Font(
        size=13,
        bold=True,
    )
    worksheet["D4"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )
    worksheet["D4"].border = THIN_BORDER

    worksheet["E4"] = len(raw_data)
    worksheet["E4"].font = Font(
        size=18,
        bold=True,
        color="1F4E78",
    )
    worksheet["E4"].fill = IQI_FILL
    worksheet["E4"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )
    worksheet["E4"].border = THIN_BORDER

    worksheet.row_dimensions[4].height = 34

    # Ομαδοποίηση των A6 και B6 για να χωράει ολόκληρη η φράση.
    worksheet.merge_cells("A6:B6")

    worksheet["A6"] = "Τελευταία ενημέρωση"
    worksheet["A6"].font = Font(
        bold=True,
    )
    worksheet["A6"].alignment = Alignment(
        horizontal="left",
        vertical="center",
    )

    # Η ημερομηνία μεταφέρεται στο C6.
    worksheet["C6"] = datetime.now()
    worksheet["C6"].number_format = "dd/mm/yyyy hh:mm"
    worksheet["C6"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    worksheet.row_dimensions[6].height = 24

    worksheet["A8"] = "Κωδικός"
    worksheet["B8"] = "Περιγραφή ενότητας"
    worksheet["C8"] = "Μέσος Όρος"

    style_table_header(
        worksheet=worksheet,
        row_number=8,
        first_column=1,
        last_column=3,
    )

    worksheet.row_dimensions[8].height = 30

    start_data_row = 9

    for row_offset, record in scores.iterrows():
        excel_row = start_data_row + row_offset

        worksheet.cell(
            row=excel_row,
            column=1,
            value=record["Section"],
        )

        worksheet.cell(
            row=excel_row,
            column=2,
            value=record["Description"],
        )

        worksheet.cell(
            row=excel_row,
            column=3,
            value=excel_safe_value(record["Mean"]),
        )

        worksheet.cell(
            row=excel_row,
            column=1,
        ).font = Font(
            bold=True,
            size=12,
        )

        worksheet.cell(
            row=excel_row,
            column=1,
        ).alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        worksheet.cell(
            row=excel_row,
            column=2,
        ).alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True,
        )

        worksheet.cell(
            row=excel_row,
            column=3,
        ).alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        worksheet.cell(
            row=excel_row,
            column=3,
        ).number_format = "0.00"

        worksheet.row_dimensions[excel_row].height = 34

    end_data_row = (
        start_data_row + len(scores) - 1
        if len(scores)
        else start_data_row
    )

    apply_table_borders(
        worksheet=worksheet,
        first_row=8,
        last_row=end_data_row,
        first_column=1,
        last_column=3,
    )

    if len(scores):
        worksheet.conditional_formatting.add(
            f"C{start_data_row}:C{end_data_row}",
            ColorScaleRule(
                start_type="num",
                start_value=1,
                start_color="F8696B",
                mid_type="num",
                mid_value=3,
                mid_color="FFEB84",
                end_type="num",
                end_value=5,
                end_color="63BE7B",
            ),
        )

        chart = BarChart()

        chart.type = "col"
        chart.grouping = "clustered"
        chart.overlap = 0
        chart.gapWidth = 65

        chart.style = 10
        chart.title = "Αξιολόγηση ανά Ενότητα"

        chart.x_axis.title = None

        # Καταργείται ο τίτλος του κατακόρυφου άξονα.
        chart.y_axis.title = None

        chart.height = 13
        chart.width = 25

        chart.legend = None

        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 5
        chart.y_axis.majorUnit = 1
        chart.y_axis.delete = False

        chart.x_axis.delete = False
        chart.x_axis.tickLblPos = "nextTo"
        chart.x_axis.tickLblSkip = 1
        chart.x_axis.tickMarkSkip = 1
        chart.x_axis.majorTickMark = "out"

        data_reference = Reference(
            worksheet,
            min_col=3,
            min_row=start_data_row,
            max_row=end_data_row,
        )

        category_reference = Reference(
            worksheet,
            min_col=1,
            min_row=start_data_row,
            max_row=end_data_row,
        )

        chart.add_data(
            data_reference,
            titles_from_data=False,
            from_rows=False,
        )

        chart.set_categories(
            category_reference
        )

        chart.dLbls = DataLabelList()
        chart.dLbls.showVal = True
        chart.dLbls.showCatName = False
        chart.dLbls.showSerName = False
        chart.dLbls.showLegendKey = False
        chart.dLbls.position = "outEnd"
        chart.dLbls.numFmt = "0.00"

        worksheet.add_chart(
            chart,
            "E7",
        )

    # Οι στήλες Α και Δ μεγαλώνουν ώστε να χωρούν οι φράσεις.
    worksheet.column_dimensions["A"].width = 26
    worksheet.column_dimensions["B"].width = 54
    worksheet.column_dimensions["C"].width = 21
    worksheet.column_dimensions["D"].width = 23

    for column_letter in [
        "E",
        "F",
        "G",
        "H",
        "I",
        "J",
        "K",
        "L",
        "M",
    ]:
        worksheet.column_dimensions[
            column_letter
        ].width = 14

    worksheet.auto_filter.ref = (
        f"A8:C{end_data_row}"
        if len(scores)
        else "A8:C8"
    )


def create_section_scores_sheet(
    workbook,
    scores,
):
    """
    Δημιουργεί αναλυτικό φύλλο μέσων όρων.
    """

    worksheet = workbook.create_sheet(
        title="Section_Scores"
    )

    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_view.zoomScale = 90
    worksheet.freeze_panes = "A2"

    display_scores = scores.rename(
        columns={
            "Section": "Κωδικός",
            "Description": "Περιγραφή ενότητας",
            "Mean": "Μέσος Όρος",
        }
    )

    write_dataframe(
        worksheet=worksheet,
        dataframe=display_scores,
    )

    style_table_header(
        worksheet=worksheet,
        row_number=1,
        first_column=1,
        last_column=len(display_scores.columns),
    )

    worksheet.row_dimensions[1].height = 30

    if len(display_scores):
        apply_table_borders(
            worksheet=worksheet,
            first_row=1,
            last_row=len(display_scores) + 1,
            first_column=1,
            last_column=len(display_scores.columns),
        )

        for row_number in range(
            2,
            len(display_scores) + 2,
        ):
            worksheet.cell(
                row=row_number,
                column=1,
            ).alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

            worksheet.cell(
                row=row_number,
                column=2,
            ).alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=True,
            )

            worksheet.cell(
                row=row_number,
                column=3,
            ).alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

            worksheet.cell(
                row=row_number,
                column=3,
            ).number_format = "0.00"

            worksheet.row_dimensions[
                row_number
            ].height = 34

        worksheet.conditional_formatting.add(
            f"C2:C{len(display_scores) + 1}",
            ColorScaleRule(
                start_type="num",
                start_value=1,
                start_color="F8696B",
                mid_type="num",
                mid_value=3,
                mid_color="FFEB84",
                end_type="num",
                end_value=5,
                end_color="63BE7B",
            ),
        )

    worksheet.column_dimensions["A"].width = 13
    worksheet.column_dimensions["B"].width = 54
    worksheet.column_dimensions["C"].width = 18


def create_dataframe_sheet(
    workbook,
    title,
    dataframe,
):
    """
    Δημιουργεί γενικό φύλλο δεδομένων.
    """

    worksheet = workbook.create_sheet(
        title=title
    )

    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_view.zoomScale = 80

    write_dataframe(
        worksheet=worksheet,
        dataframe=dataframe,
    )

    if len(dataframe.columns):
        style_table_header(
            worksheet=worksheet,
            row_number=1,
            first_column=1,
            last_column=len(dataframe.columns),
        )

        worksheet.row_dimensions[1].height = 45

        worksheet.auto_filter.ref = (
            f"A1:"
            f"{get_column_letter(len(dataframe.columns))}"
            f"{len(dataframe) + 1}"
        )

        for column_number in range(
            1,
            len(dataframe.columns) + 1,
        ):
            worksheet.cell(
                row=1,
                column=column_number,
            ).alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

    auto_fit_columns(
        worksheet,
        minimum_width=14,
        maximum_width=65,
    )


def create_master():
    """
    Δημιουργεί το αρχείο ΣΑΠ-ΦΘ_MASTER.xlsx.
    """

    raw_data = load_data()
    scored_data = build_scored_dataset()
    metadata = build_metadata()

    scores = prepare_section_scores(
        section_means()
    )

    iqi_value = calculate_iqi()

    workbook = Workbook()

    create_dashboard(
        workbook=workbook,
        raw_data=raw_data,
        scores=scores,
        iqi_value=iqi_value,
    )

    create_section_scores_sheet(
        workbook=workbook,
        scores=scores,
    )

    create_dataframe_sheet(
        workbook=workbook,
        title="Raw_Data",
        dataframe=raw_data,
    )

    create_dataframe_sheet(
        workbook=workbook,
        title="Scored_Data",
        dataframe=scored_data,
    )

    create_dataframe_sheet(
        workbook=workbook,
        title="Metadata",
        dataframe=metadata,
    )

    MASTER_FILE.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook.save(MASTER_FILE)

    print(
        f"MASTER.xlsx δημιουργήθηκε: {MASTER_FILE}"
    )


if __name__ == "__main__":
    create_master()