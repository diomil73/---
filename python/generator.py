"""
ΣΑΠ-ΦΘ
MASTER Excel Generator
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font

from config import (
    PROJECT_FULL_NAME,
    HOSPITAL,
    ORGANIZATION,
    MASTER_FILE
)

SHEETS = [
    "00_ΑΡΧΙΚΗ",
    "01_ΡΥΘΜΙΣΕΙΣ",
    "02_ΑΠΑΝΤΗΣΕΙΣ",
    "03_ΚΑΘΑΡΑ_ΔΕΔΟΜΕΝΑ",
    "04_ΥΠΟΛΟΓΙΣΜΟΙ",
    "05_DASHBOARD",
    "06_ΣΧΟΛΙΑ",
    "07_ΑΝΑΦΟΡΑ",
    "08_ΛΕΞΙΚΟ_ΟΡΩΝ",
    "09_CHANGELOG",
]

def create_master():

    wb = Workbook()

    wb.remove(wb.active)

    for sheet in SHEETS:

        ws = wb.create_sheet(sheet)

        ws["A1"] = sheet
        ws["A1"].font = Font(size=14, bold=True)

    home = wb["00_ΑΡΧΙΚΗ"]

    home["A3"] = PROJECT_FULL_NAME
    home["A5"] = "Φορέας"
    home["B5"] = ORGANIZATION
    home["A6"] = "Νοσοκομείο"
    home["B6"] = HOSPITAL

    MASTER_FILE.parent.mkdir(parents=True, exist_ok=True)

    wb.save(MASTER_FILE)

    print(f"MASTER δημιουργήθηκε:\n{MASTER_FILE}")

if __name__ == "__main__":
    create_master()
