"""
ΣΑΠ-ΦΘ
Αναλυτική αξιολόγηση και γραφήματα ανά ερώτημα.
"""

from pathlib import Path
import re

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)

from import_data import load_data
from metadata_engine import build_metadata
from scoring_engine import build_scored_dataset


PROJECT_ROOT = Path(__file__).resolve().parent.parent

OUTPUT_FILE = (
    PROJECT_ROOT
    / "excel"
    / "QUESTION_LEVEL_ANALYTICS.xlsx"
)


SECTION_ORDER = [
    "Δημογραφικά",
    "Α",
    "Β",
    "Γ",
    "Δ",
    "Ε",
    "ΣΤ",
    "Ζ",
    "Η",
    "Θ",
]


SECTION_TITLES = {
    "Δημογραφικά": "Γενικά και δημογραφικά στοιχεία",
    "Α": "Υποδοχή και ενημέρωση",
    "Β": "Οργάνωση πρακτικής άσκησης",
    "Γ": "Κλινική εκπαίδευση και εμπειρία",
    "Δ": "Υποδομές και εξοπλισμός",
    "Ε": "Συνεργασία με επαγγελματίες υγείας",
    "ΣΤ": "Καθοδήγηση και υποστήριξη",
    "Ζ": "Ηγεσία και επαγγελματισμός",
    "Η": "Κλίμα συνεργασίας και ασφάλειας",
    "Θ": "Χώροι και συνθήκες πρακτικής",
}


TITLE_FILL = PatternFill(
    fill_type="solid",
    fgColor="1F4E78",
)

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="5B9BD5",
)

LABEL_FILL = PatternFill(
    fill_type="solid",
    fgColor="D9EAF7",
)

VALUE_FILL = PatternFill(
    fill_type="solid",
    fgColor="E2F0D9",
)

WHITE_FONT = Font(
    color="FFFFFF",
    bold=True,
)

THIN_BORDER = Border(
    left=Side(
        style="thin",
        color="B7B7B7",
    ),
    right=Side(
        style="thin",
        color="B7B7B7",
    ),
    top=Side(
        style="thin",
        color="B7B7B7",
    ),
    bottom=Side(
        style="thin",
        color="B7B7B7",
    ),
)


def clean_text(value):
    """
    Καθαρίζει περιττά κενά και εξωτερικές αγκύλες.
    """

    if value is None:
        return ""

    text = re.sub(
        r"\s+",
        " ",
        str(value).strip(),
    )

    return text.strip("[]").strip()


def excel_safe_value(value):
    """
    Μετατρέπει κενές τιμές pandas σε κενό Excel.
    """

    if value is None:
        return None

    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass

    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()

    return value


def section_order_value(section):
    """
    Επιστρέφει τη σειρά εμφάνισης της ενότητας.
    """

    section = clean_text(section)

    if section in SECTION_ORDER:
        return SECTION_ORDER.index(section)

    return len(SECTION_ORDER)


def normalize_metadata(metadata):
    """
    Κανονικοποιεί το metadata χωρίς να αλλάζει
    τα πραγματικά ονόματα των ερωτήσεων.
    """

    required_columns = [
        "Question",
        "Section",
        "Type",
    ]

    missing_columns = [
        column
        for column in required_columns
        if column not in metadata.columns
    ]

    if missing_columns:
        raise ValueError(
            "Λείπουν από το metadata οι στήλες: "
            + ", ".join(missing_columns)
        )

    normalized = metadata.copy()

    normalized["Question"] = (
        normalized["Question"]
        .astype(str)
    )

    normalized["Section"] = (
        normalized["Section"]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    normalized["Type"] = (
        normalized["Type"]
        .fillna("")
        .astype(str)
        .str.strip()
        .str.upper()
    )

    normalized["_order"] = (
        normalized["Section"]
        .apply(section_order_value)
    )

    normalized = normalized.sort_values(
        by=[
            "_order",
            "Question",
        ]
    )

    return normalized.reset_index(drop=True)


def should_exclude_question(
    question,
    question_type,
):
    """
    Εξαιρεί πεδία που δεν αποτελούν
    αριθμητική ή κατηγορηματική αξιολόγηση.
    """

    question_clean = clean_text(
        question
    ).lower()

    question_type = clean_text(
        question_type
    ).upper()

    if question_type == "TEXT":
        return True

    excluded_questions = {
        "timestamp",
        "χρονική σήμανση",
    }

    return question_clean in excluded_questions


def get_valid_raw_values(series):
    """
    Επιστρέφει τις μη κενές αρχικές απαντήσεις.
    """

    valid_values = (
        series
        .dropna()
        .astype(str)
        .str.strip()
    )

    return valid_values.loc[
        valid_values != ""
    ]


def get_numeric_values(
    question,
    raw_data,
    scored_data,
):
    """
    Αναζητά πρώτα αριθμητικές βαθμολογίες
    στο Scored_Data και μετά στο Raw_Data.
    """

    if question in scored_data.columns:
        scored_values = pd.to_numeric(
            scored_data[question],
            errors="coerce",
        )

        if scored_values.notna().any():
            return scored_values

    return pd.to_numeric(
        raw_data[question],
        errors="coerce",
    )


def classify_question(
    question_type,
    raw_series,
    numeric_series,
):
    """
    Κατατάσσει μία ερώτηση σε βαθμολογική,
    αριθμητική ή κατηγορηματική.
    """

    normalized_type = clean_text(
        question_type
    ).upper()

    if normalized_type == "LIKERT":
        return "Βαθμολογική"

    if normalized_type == "NPS":
        return "Αριθμητική"

    valid_raw_count = len(
        get_valid_raw_values(raw_series)
    )

    numeric_count = int(
        numeric_series.notna().sum()
    )

    if (
        valid_raw_count > 0
        and numeric_count / valid_raw_count >= 0.70
    ):
        return "Αριθμητική"

    return "Κατηγορηματική"


def create_likert_distribution(
    numeric_series,
):
    """
    Δημιουργεί πλήρη κατανομή βαθμολογιών 1–5.
    """

    valid_values = pd.to_numeric(
        numeric_series,
        errors="coerce",
    ).dropna()

    total = len(valid_values)
    records = []

    for category in [1, 2, 3, 4, 5]:
        count = int(
            (valid_values == category).sum()
        )

        percentage = (
            count / total * 100
            if total
            else 0
        )

        records.append(
            {
                "Κατηγορία": str(category),
                "Πλήθος": count,
                "Ποσοστό": round(
                    percentage,
                    2,
                ),
            }
        )

    return pd.DataFrame(records)


def create_numeric_distribution(
    numeric_series,
):
    """
    Δημιουργεί κατανομή αριθμητικών απαντήσεων.
    """

    valid_values = pd.to_numeric(
        numeric_series,
        errors="coerce",
    ).dropna()

    counts = (
        valid_values
        .value_counts()
        .sort_index()
    )

    total = int(
        counts.sum()
    )

    records = []

    for category, count in counts.items():
        category_number = float(category)

        if category_number.is_integer():
            category_label = str(
                int(category_number)
            )
        else:
            category_label = str(
                round(category_number, 2)
            )

        percentage = (
            int(count) / total * 100
            if total
            else 0
        )

        records.append(
            {
                "Κατηγορία": category_label,
                "Πλήθος": int(count),
                "Ποσοστό": round(
                    percentage,
                    2,
                ),
            }
        )

    return pd.DataFrame(records)


def create_categorical_distribution(
    raw_series,
):
    """
    Δημιουργεί κατανομή κατηγορηματικών απαντήσεων.
    """

    valid_values = get_valid_raw_values(
        raw_series
    )

    counts = valid_values.value_counts()

    total = int(
        counts.sum()
    )

    records = []

    for category, count in counts.items():
        percentage = (
            int(count) / total * 100
            if total
            else 0
        )

        records.append(
            {
                "Κατηγορία": clean_text(
                    category
                ),
                "Πλήθος": int(count),
                "Ποσοστό": round(
                    percentage,
                    2,
                ),
            }
        )

    return pd.DataFrame(records)


def calculate_question_result(
    question,
    section,
    question_type,
    raw_data,
    scored_data,
):
    """
    Υπολογίζει τα αναλυτικά αποτελέσματα
    ενός ερωτήματος.
    """

    raw_series = raw_data[question]

    numeric_series = get_numeric_values(
        question=question,
        raw_data=raw_data,
        scored_data=scored_data,
    )

    question_class = classify_question(
        question_type=question_type,
        raw_series=raw_series,
        numeric_series=numeric_series,
    )

    total_responses = len(raw_series)

    valid_raw = get_valid_raw_values(
        raw_series
    )

    raw_valid_count = len(valid_raw)

    mean_value = None
    median_value = None
    minimum_value = None
    maximum_value = None
    mode_value = None
    non_scored_count = 0

    if question_class in {
        "Βαθμολογική",
        "Αριθμητική",
    }:
        valid_numeric = pd.to_numeric(
            numeric_series,
            errors="coerce",
        ).dropna()

        valid_count = len(valid_numeric)

        non_scored_count = max(
            raw_valid_count - valid_count,
            0,
        )

        if valid_count:
            mean_value = round(
                float(valid_numeric.mean()),
                2,
            )

            median_value = round(
                float(valid_numeric.median()),
                2,
            )

            minimum_value = round(
                float(valid_numeric.min()),
                2,
            )

            maximum_value = round(
                float(valid_numeric.max()),
                2,
            )

            modes = valid_numeric.mode()

            if not modes.empty:
                mode_number = float(
                    modes.iloc[0]
                )

                if mode_number.is_integer():
                    mode_value = int(
                        mode_number
                    )
                else:
                    mode_value = round(
                        mode_number,
                        2,
                    )

        if question_class == "Βαθμολογική":
            distribution = (
                create_likert_distribution(
                    numeric_series
                )
            )
        else:
            distribution = (
                create_numeric_distribution(
                    numeric_series
                )
            )

    else:
        valid_count = raw_valid_count

        distribution = (
            create_categorical_distribution(
                raw_series
            )
        )

        if not distribution.empty:
            mode_value = distribution.iloc[0][
                "Κατηγορία"
            ]

    missing_count = (
        total_responses
        - raw_valid_count
    )

    return {
        "question": clean_text(question),
        "section": clean_text(section),
        "section_title": SECTION_TITLES.get(
            clean_text(section),
            "",
        ),
        "question_type": clean_text(
            question_type
        ),
        "question_class": question_class,
        "total_responses": total_responses,
        "valid_responses": valid_count,
        "missing_responses": missing_count,
        "non_scored_responses": non_scored_count,
        "mean": mean_value,
        "median": median_value,
        "minimum": minimum_value,
        "maximum": maximum_value,
        "mode": mode_value,
        "distribution": distribution,
    }


def collect_question_results(
    raw_data,
    scored_data,
    metadata,
):
    """
    Συλλέγει όλες τις αριθμητικές,
    βαθμολογικές και κατηγορηματικές ερωτήσεις.
    """

    normalized_metadata = normalize_metadata(
        metadata
    )

    results = []

    examined_count = 0
    excluded_count = 0
    empty_count = 0

    for _, metadata_row in (
        normalized_metadata.iterrows()
    ):
        question = metadata_row["Question"]
        section = metadata_row["Section"]
        question_type = metadata_row["Type"]

        if question not in raw_data.columns:
            continue

        examined_count += 1

        if should_exclude_question(
            question=question,
            question_type=question_type,
        ):
            excluded_count += 1
            continue

        result = calculate_question_result(
            question=question,
            section=section,
            question_type=question_type,
            raw_data=raw_data,
            scored_data=scored_data,
        )

        if result["distribution"].empty:
            empty_count += 1
            continue

        results.append(result)

    print()
    print(
        f"Ερωτήματα που εξετάστηκαν: "
        f"{examined_count}"
    )

    print(
        f"Ερωτήματα που εξαιρέθηκαν: "
        f"{excluded_count}"
    )

    print(
        "Ερωτήματα χωρίς διαθέσιμη κατανομή: "
        f"{empty_count}"
    )

    print(
        f"Ερωτήματα που θα παρουσιαστούν: "
        f"{len(results)}"
    )

    return results


def style_header_row(
    worksheet,
    row_number,
    first_column,
    last_column,
):
    """
    Μορφοποιεί γραμμή επικεφαλίδων.
    """

    for column_number in range(
        first_column,
        last_column + 1,
    ):
        cell = worksheet.cell(
            row=row_number,
            column=column_number,
        )

        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )


def create_index_sheet(
    worksheet,
    results,
    block_starts,
):
    """
    Δημιουργεί το ευρετήριο όλων των ερωτημάτων.
    """

    worksheet.title = "Question_Index"

    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_view.zoomScale = 85

    headers = [
        "Α/Α",
        "Ενότητα",
        "Περιγραφή ενότητας",
        "Ερώτημα",
        "Τύπος ανάλυσης",
        "Έγκυρες απαντήσεις",
        "Μη βαθμολογήσιμες",
        "Μέσος όρος",
        "Επικρατέστερη απάντηση",
        "Μετάβαση",
    ]

    for column_number, header in enumerate(
        headers,
        start=1,
    ):
        worksheet.cell(
            row=1,
            column=column_number,
            value=header,
        )

    style_header_row(
        worksheet=worksheet,
        row_number=1,
        first_column=1,
        last_column=len(headers),
    )

    for index, result in enumerate(
        results,
        start=1,
    ):
        row_number = index + 1

        values = [
            index,
            result["section"],
            result["section_title"],
            result["question"],
            result["question_class"],
            result["valid_responses"],
            result["non_scored_responses"],
            result["mean"],
            result["mode"],
            "Άνοιγμα γραφήματος",
        ]

        for column_number, value in enumerate(
            values,
            start=1,
        ):
            cell = worksheet.cell(
                row=row_number,
                column=column_number,
                value=excel_safe_value(value),
            )

            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=(
                    column_number
                    in [3, 4, 9, 10]
                ),
            )

        worksheet.cell(
            row=row_number,
            column=8,
        ).number_format = "0.00"

        link_cell = worksheet.cell(
            row=row_number,
            column=10,
        )

        link_cell.hyperlink = (
            f"#'Question_Charts'!"
            f"A{block_starts[index - 1]}"
        )

        link_cell.style = "Hyperlink"

        worksheet.row_dimensions[
            row_number
        ].height = 45

    worksheet.column_dimensions["A"].width = 8
    worksheet.column_dimensions["B"].width = 13
    worksheet.column_dimensions["C"].width = 38
    worksheet.column_dimensions["D"].width = 75
    worksheet.column_dimensions["E"].width = 20
    worksheet.column_dimensions["F"].width = 18
    worksheet.column_dimensions["G"].width = 19
    worksheet.column_dimensions["H"].width = 15
    worksheet.column_dimensions["I"].width = 30
    worksheet.column_dimensions["J"].width = 22

    if results:
        worksheet.conditional_formatting.add(
            f"H2:H{len(results) + 1}",
            ColorScaleRule(
                start_type="num",
                start_value=1,
                start_color="F8696B",
                mid_type="num",
                mid_value=3,
                mid_color="FFEB84",
                end_type="num",
                end_value=5,
                end_color="63BE7B",
            ),
        )

        worksheet.auto_filter.ref = (
            f"A1:J{len(results) + 1}"
        )


def add_question_block(
    worksheet,
    result,
    question_number,
    start_row,
):
    """
    Δημιουργεί πλήρες μπλοκ αποτελεσμάτων
    και γράφημα για μία ερώτηση.
    """

    worksheet.merge_cells(
        start_row=start_row,
        start_column=1,
        end_row=start_row,
        end_column=10,
    )

    title_cell = worksheet.cell(
        row=start_row,
        column=1,
        value=(
            f"{question_number}. "
            f"{result['question']}"
        ),
    )

    title_cell.fill = TITLE_FILL

    title_cell.font = Font(
        color="FFFFFF",
        bold=True,
        size=12,
    )

    title_cell.alignment = Alignment(
        vertical="center",
        wrap_text=True,
    )

    worksheet.row_dimensions[
        start_row
    ].height = 48

    summary_rows = [
        (
            "Ενότητα",
            (
                f"{result['section']} – "
                f"{result['section_title']}"
            ),
            "Τύπος",
            result["question_class"],
        ),
        (
            "Έγκυρες απαντήσεις",
            result["valid_responses"],
            "Κενές απαντήσεις",
            result["missing_responses"],
        ),
        (
            "Μη βαθμολογήσιμες",
            result["non_scored_responses"],
            "Αρχικός τύπος",
            result["question_type"],
        ),
        (
            "Μέσος όρος",
            result["mean"],
            "Διάμεσος",
            result["median"],
        ),
        (
            "Ελάχιστο",
            result["minimum"],
            "Μέγιστο",
            result["maximum"],
        ),
        (
            "Επικρατέστερη απάντηση",
            result["mode"],
            "",
            "",
        ),
    ]

    for row_offset, row_values in enumerate(
        summary_rows,
        start=1,
    ):
        current_row = start_row + row_offset

        worksheet.cell(
            row=current_row,
            column=1,
            value=row_values[0],
        )

        worksheet.cell(
            row=current_row,
            column=2,
            value=excel_safe_value(
                row_values[1]
            ),
        )

        worksheet.cell(
            row=current_row,
            column=4,
            value=row_values[2],
        )

        worksheet.cell(
            row=current_row,
            column=5,
            value=excel_safe_value(
                row_values[3]
            ),
        )

        for column_number in [1, 4]:
            cell = worksheet.cell(
                row=current_row,
                column=column_number,
            )

            cell.fill = LABEL_FILL
            cell.font = Font(
                bold=True
            )
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=True,
            )

        for column_number in [2, 5]:
            cell = worksheet.cell(
                row=current_row,
                column=column_number,
            )

            cell.fill = VALUE_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

    worksheet.cell(
        row=start_row + 4,
        column=2,
    ).number_format = "0.00"

    worksheet.cell(
        row=start_row + 4,
        column=5,
    ).number_format = "0.00"

    worksheet.cell(
        row=start_row + 5,
        column=2,
    ).number_format = "0.00"

    worksheet.cell(
        row=start_row + 5,
        column=5,
    ).number_format = "0.00"

    table_header_row = start_row + 8

    table_headers = [
        "Κατηγορία",
        "Πλήθος",
        "Ποσοστό %",
    ]

    for column_number, header in enumerate(
        table_headers,
        start=1,
    ):
        worksheet.cell(
            row=table_header_row,
            column=column_number,
            value=header,
        )

    style_header_row(
        worksheet=worksheet,
        row_number=table_header_row,
        first_column=1,
        last_column=3,
    )

    distribution = result["distribution"]

    first_data_row = table_header_row + 1

    for row_offset, record in (
        distribution.iterrows()
    ):
        current_row = (
            first_data_row + row_offset
        )

        worksheet.cell(
            row=current_row,
            column=1,
            value=excel_safe_value(
                record["Κατηγορία"]
            ),
        )

        worksheet.cell(
            row=current_row,
            column=2,
            value=int(
                record["Πλήθος"]
            ),
        )

        worksheet.cell(
            row=current_row,
            column=3,
            value=float(
                record["Ποσοστό"]
            ),
        )

        worksheet.cell(
            row=current_row,
            column=3,
        ).number_format = "0.00"

        for column_number in range(
            1,
            4,
        ):
            cell = worksheet.cell(
                row=current_row,
                column=column_number,
            )

            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

    last_data_row = (
        first_data_row
        + len(distribution)
        - 1
    )

    chart = BarChart()

    chart.type = "col"
    chart.grouping = "clustered"
    chart.overlap = 0
    chart.gapWidth = 65

    chart.title = "Κατανομή απαντήσεων"
    chart.legend = None

    chart.x_axis.title = None
    chart.y_axis.title = None

    chart.height = 8.5
    chart.width = 15

    chart.x_axis.delete = False
    chart.x_axis.tickLblPos = "nextTo"

    chart.y_axis.scaling.min = 0
    chart.y_axis.majorUnit = 1

    data_reference = Reference(
        worksheet,
        min_col=2,
        min_row=table_header_row,
        max_row=last_data_row,
    )

    category_reference = Reference(
        worksheet,
        min_col=1,
        min_row=first_data_row,
        max_row=last_data_row,
    )

    chart.add_data(
        data_reference,
        titles_from_data=True,
    )

    chart.set_categories(
        category_reference
    )

    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = True
    chart.dLbls.showCatName = False
    chart.dLbls.showSerName = False
    chart.dLbls.showLegendKey = False
    chart.dLbls.dLblPos = "outEnd"

    worksheet.add_chart(
        chart,
        f"E{table_header_row}",
    )

    block_height = max(
        21,
        11 + len(distribution),
    )

    return start_row + block_height


def create_charts_sheet(
    workbook,
    results,
):
    """
    Δημιουργεί το φύλλο όλων των γραφημάτων.
    """

    worksheet = workbook.create_sheet(
        title="Question_Charts"
    )

    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_view.zoomScale = 80

    worksheet.column_dimensions["A"].width = 34
    worksheet.column_dimensions["B"].width = 20
    worksheet.column_dimensions["C"].width = 5
    worksheet.column_dimensions["D"].width = 25
    worksheet.column_dimensions["E"].width = 20

    for column_letter in [
        "F",
        "G",
        "H",
        "I",
        "J",
    ]:
        worksheet.column_dimensions[
            column_letter
        ].width = 14

    block_starts = []
    current_row = 1

    for question_number, result in enumerate(
        results,
        start=1,
    ):
        block_starts.append(
            current_row
        )

        next_row = add_question_block(
            worksheet=worksheet,
            result=result,
            question_number=question_number,
            start_row=current_row,
        )

        current_row = next_row + 2

    return block_starts


def create_question_level_analytics():
    """
    Δημιουργεί το QUESTION_LEVEL_ANALYTICS.xlsx.
    """

    raw_data = load_data()
    scored_data = build_scored_dataset()
    metadata = build_metadata()

    results = collect_question_results(
        raw_data=raw_data,
        scored_data=scored_data,
        metadata=metadata,
    )

    workbook = Workbook()

    index_worksheet = workbook.active

    block_starts = create_charts_sheet(
        workbook=workbook,
        results=results,
    )

    create_index_sheet(
        worksheet=index_worksheet,
        results=results,
        block_starts=block_starts,
    )

    workbook.active = 0

    OUTPUT_FILE.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook.save(
        OUTPUT_FILE
    )

    numeric_count = sum(
        1
        for result in results
        if result["question_class"]
        in {
            "Βαθμολογική",
            "Αριθμητική",
        }
    )

    categorical_count = sum(
        1
        for result in results
        if result["question_class"]
        == "Κατηγορηματική"
    )

    print()

    print(
        "Αναλυτική αξιολόγηση ανά ερώτημα "
        f"δημιουργήθηκε: {OUTPUT_FILE}"
    )

    print(
        "Ερωτήματα με αριθμητική αξιολόγηση: "
        f"{numeric_count}"
    )

    print(
        "Κατηγορηματικά ερωτήματα: "
        f"{categorical_count}"
    )

    print(
        "Συνολικά ερωτήματα με γράφημα: "
        f"{len(results)}"
    )


if __name__ == "__main__":
    create_question_level_analytics()