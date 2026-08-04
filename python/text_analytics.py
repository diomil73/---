"""
ΣΑΠ-ΦΘ
Open Text Analytics
"""

from collections import Counter
from pathlib import Path
import re
import unicodedata

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from import_data import load_data
from metadata_engine import build_metadata


PROJECT_ROOT = Path(__file__).resolve().parent.parent

OUTPUT_FILE = (
    PROJECT_ROOT
    / "excel"
    / "TEXT_ANALYTICS.xlsx"
)


STOPWORDS = {
    "και",
    "να",
    "το",
    "τη",
    "την",
    "της",
    "του",
    "των",
    "τα",
    "σε",
    "στο",
    "στη",
    "στην",
    "στον",
    "με",
    "για",
    "που",
    "από",
    "ως",
    "ή",
    "είναι",
    "ήταν",
    "έχει",
    "έχουν",
    "θα",
    "δεν",
    "μια",
    "ένα",
    "ένας",
    "οι",
    "ο",
    "η",
    "μας",
    "μου",
    "σας",
    "τους",
    "τις",
    "αυτό",
    "αυτή",
    "αυτά",
    "πολύ",
    "πιο",
    "κατά",
    "στης",
    "στις",
    "στους",
    "μέσα",
    "επίσης",
    "ότι",
    "όταν",
    "όπως",
    "μπορεί",
    "ήθελα",
    "ήθελαμε",
    "υπήρχε",
    "υπήρχαν",
}


HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="1F4E78",
)

HEADER_FONT = Font(
    color="FFFFFF",
    bold=True,
)


def clean_question(question):
    """
    Αφαιρεί περιττά κενά και αγκύλες.
    """

    return str(question).strip().strip("[]").strip()


def normalize_text(value):
    """
    Μετατρέπει το κείμενο σε πεζά,
    αφαιρεί τόνους και ειδικούς χαρακτήρες.
    """

    text = str(value).strip().lower()

    text = unicodedata.normalize(
        "NFD",
        text,
    )

    text = "".join(
        character
        for character in text
        if unicodedata.category(character) != "Mn"
    )

    text = re.sub(
        r"[^a-zA-Zα-ωΑ-Ωάέήίόύώϊϋΐΰ0-9\s]",
        " ",
        text,
    )

    text = re.sub(
        r"\s+",
        " ",
        text,
    ).strip()

    return text


def tokenize(value):
    """
    Χωρίζει ένα σχόλιο σε χρήσιμες λέξεις.
    """

    normalized = normalize_text(value)

    words = normalized.split()

    return [
        word
        for word in words
        if (
            len(word) >= 3
            and word not in STOPWORDS
            and not word.isdigit()
        )
    ]


def build_comments_table(
    raw_data,
    metadata,
):
    """
    Δημιουργεί πίνακα με όλες τις ανοιχτές απαντήσεις.
    """

    records = []

    text_questions = metadata.loc[
        metadata["Type"] == "TEXT",
        "Question",
    ].tolist()

    for question in text_questions:
        if question not in raw_data.columns:
            continue

        clean_name = clean_question(question)

        for row_index, value in raw_data[question].items():
            if pd.isna(value):
                continue

            text_value = str(value).strip()

            if not text_value:
                continue

            records.append(
                {
                    "Γραμμή δεδομένων": row_index + 2,
                    "Ερώτηση": clean_name,
                    "Απάντηση": text_value,
                    "Αριθμός λέξεων": len(
                        text_value.split()
                    ),
                    "Αριθμός χαρακτήρων": len(
                        text_value
                    ),
                }
            )

    return pd.DataFrame(
        records,
        columns=[
            "Γραμμή δεδομένων",
            "Ερώτηση",
            "Απάντηση",
            "Αριθμός λέξεων",
            "Αριθμός χαρακτήρων",
        ],
    )


def build_question_summary(
    raw_data,
    metadata,
):
    """
    Δημιουργεί σύνοψη ανά ανοιχτή ερώτηση.
    """

    records = []

    text_questions = metadata.loc[
        metadata["Type"] == "TEXT",
        "Question",
    ].tolist()

    total_students = len(raw_data)

    for question in text_questions:
        if question not in raw_data.columns:
            continue

        valid_values = (
            raw_data[question]
            .dropna()
            .astype(str)
            .str.strip()
        )

        valid_values = valid_values.loc[
            valid_values != ""
        ]

        response_count = len(valid_values)

        blank_count = (
            total_students
            - response_count
        )

        word_counts = valid_values.apply(
            lambda value: len(value.split())
        )

        character_counts = valid_values.apply(
            len
        )

        response_percentage = (
            round(
                response_count
                / total_students
                * 100,
                2,
            )
            if total_students
            else 0
        )

        average_words = (
            round(
                float(word_counts.mean()),
                2,
            )
            if response_count
            else 0
        )

        average_characters = (
            round(
                float(character_counts.mean()),
                2,
            )
            if response_count
            else 0
        )

        records.append(
            {
                "Ερώτηση": clean_question(question),
                "Σύνολο φοιτητών": total_students,
                "Απαντήσεις": response_count,
                "Κενές απαντήσεις": blank_count,
                "Ποσοστό απάντησης %": (
                    response_percentage
                ),
                "Μέσος αριθμός λέξεων": (
                    average_words
                ),
                "Μέσος αριθμός χαρακτήρων": (
                    average_characters
                ),
            }
        )

    return pd.DataFrame(records)


def build_keyword_table(
    comments,
):
    """
    Υπολογίζει συχνές λέξεις ανά ερώτηση.
    """

    records = []

    if comments.empty:
        return pd.DataFrame(
            columns=[
                "Ερώτηση",
                "Λέξη",
                "Συχνότητα",
            ]
        )

    for question, group in comments.groupby(
        "Ερώτηση"
    ):
        word_counter = Counter()

        for response in group["Απάντηση"]:
            word_counter.update(
                tokenize(response)
            )

        for word, frequency in word_counter.most_common(
            20
        ):
            records.append(
                {
                    "Ερώτηση": question,
                    "Λέξη": word,
                    "Συχνότητα": frequency,
                }
            )

    return pd.DataFrame(records)


def build_all_keywords_table(
    comments,
):
    """
    Υπολογίζει τις συχνότερες λέξεις συνολικά.
    """

    counter = Counter()

    if not comments.empty:
        for response in comments["Απάντηση"]:
            counter.update(
                tokenize(response)
            )

    records = [
        {
            "Λέξη": word,
            "Συχνότητα": frequency,
        }
        for word, frequency in counter.most_common(
            50
        )
    ]

    return pd.DataFrame(
        records,
        columns=[
            "Λέξη",
            "Συχνότητα",
        ],
    )


def style_worksheet(
    worksheet,
    wrap_columns=None,
    maximum_width=60,
):
    """
    Μορφοποιεί ένα φύλλο του Excel.
    """

    wrap_columns = wrap_columns or []

    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False
    worksheet.auto_filter.ref = worksheet.dimensions

    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    worksheet.row_dimensions[1].height = 35

    for column_number in range(
        1,
        worksheet.max_column + 1,
    ):
        column_letter = get_column_letter(
            column_number
        )

        maximum_length = 0

        for row_number in range(
            1,
            worksheet.max_row + 1,
        ):
            cell = worksheet.cell(
                row=row_number,
                column=column_number,
            )

            if cell.value is not None:
                maximum_length = max(
                    maximum_length,
                    len(str(cell.value)),
                )

            if column_number in wrap_columns:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

        worksheet.column_dimensions[
            column_letter
        ].width = min(
            max(maximum_length + 3, 12),
            maximum_width,
        )

    for row_number in range(
        2,
        worksheet.max_row + 1,
    ):
        worksheet.row_dimensions[
            row_number
        ].height = 40


def create_text_analytics():
    """
    Δημιουργεί το TEXT_ANALYTICS.xlsx.
    """

    raw_data = load_data()
    metadata = build_metadata()

    comments = build_comments_table(
        raw_data=raw_data,
        metadata=metadata,
    )

    question_summary = build_question_summary(
        raw_data=raw_data,
        metadata=metadata,
    )

    keywords_by_question = build_keyword_table(
        comments
    )

    all_keywords = build_all_keywords_table(
        comments
    )

    OUTPUT_FILE.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    with pd.ExcelWriter(
        OUTPUT_FILE,
        engine="openpyxl",
    ) as writer:
        question_summary.to_excel(
            writer,
            sheet_name="Summary",
            index=False,
        )

        comments.to_excel(
            writer,
            sheet_name="All_Comments",
            index=False,
        )

        keywords_by_question.to_excel(
            writer,
            sheet_name="Keywords_By_Question",
            index=False,
        )

        all_keywords.to_excel(
            writer,
            sheet_name="Overall_Keywords",
            index=False,
        )

        workbook = writer.book

        style_worksheet(
            workbook["Summary"],
            maximum_width=55,
        )

        style_worksheet(
            workbook["All_Comments"],
            wrap_columns=[2, 3],
            maximum_width=75,
        )

        style_worksheet(
            workbook["Keywords_By_Question"],
            wrap_columns=[1],
            maximum_width=70,
        )

        style_worksheet(
            workbook["Overall_Keywords"],
            maximum_width=35,
        )

    print(
        f"Text analytics δημιουργήθηκε: {OUTPUT_FILE}"
    )

    print(
        "Ανοιχτές ερωτήσεις που αναλύθηκαν: "
        f"{len(question_summary)}"
    )

    print(
        "Συνολικά σχόλια που αναλύθηκαν: "
        f"{len(comments)}"
    )


if __name__ == "__main__":
    create_text_analytics()